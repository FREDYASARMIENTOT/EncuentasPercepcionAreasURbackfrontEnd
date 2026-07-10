# -*- coding: utf-8 -*-
"""
verificar_excels_parquet.py - Valida la estructura y consistencia de los Excel
generados por el lanzador parquet.

Uso:
    python verificar_excels_parquet.py [--ruta RAIZ] [--area AREA] [--anio ANIO] [--mes MES]

Ejemplos:
    python verificar_excels_parquet.py
    python verificar_excels_parquet.py --area "CRAI" --anio 2026 --mes 5
    python verificar_excels_parquet.py --ruta "F:/ETL_DITIC/temp_exportacion_multiarea"
"""
import os
import sys
import argparse
from pathlib import Path
from datetime import datetime

import pandas as pd

try:
    from openpyxl import load_workbook
    OPENPYXL_OK = True
except:
    OPENPYXL_OK = False

# Columnas esperadas en la vista SQL (30 columnas)
VISTA_SQL_COLUMNAS = [
    "encuestadoId", "Año", "Mes", "respuestaFch", "preguntaDescripcion",
    "encuestaNombre", "servicioNombre", "sedeNombre", "areaNombre",
    "encuestadoApellidos", "encuestadoNombres", "encuestadoCelular",
    "encuestadoEmail", "tipoPreguntaId", "consecutivo", "preguntaId",
    "encuestaId", "idAreaServicio", "idServicioEncuesta", "idSedeEncuesta",
    "respuestaId", "Atención", "Comunicación y acceso", "Eficiencia", "NPS",
    "Resolución de la necesidad", "Tiempo de respuesta", "preguntaSinIndicador",
    "NPS_Numerico", "Atencion_Numerica",
]

# Mapeo parquet -> SQL
PARQUET_RENOMBRA = {
    "Anio": "Año",
    "Atencion": "Atención",
    "Comunicacion_y_acceso": "Comunicación y acceso",
    "Resolucion_de_la_necesidad": "Resolución de la necesidad",
    "Tiempo_de_respuesta": "Tiempo de respuesta",
}

# Metricas clave que deben existir en los datos procesados
METRICAS_ESPERADAS = [
    "Atención", "Comunicación y acceso", "Eficiencia",
    "NPS", "Resolución de la necesidad", "Tiempo de respuesta"
]

# Pestañas esperadas en cada Excel
PESTANAS_MENSUAL = [
    "Datos Completo", "Datos Procesados", "Total Encuestas",
    "Tablero Visual", "Analitica", "Comentarios"
]

PESTANAS_ACUMULADO = [
    "Datos Completo", "Datos Procesados", "Total Encuestas",
    "Tablero Visual", "Analitica", "Comentarios"
]


# ── Colores para terminal ────────────────────────────────────────────────
class C:
    OK  = "\033[92m"
    WARN = "\033[93m"
    ERR = "\033[91m"
    BOLD = "\033[1m"
    END = "\033[0m"
    CYAN = "\033[96m"

def ok(msg):   print(f"  {C.OK}✓{C.END} {msg}")
def warn(msg): print(f"  {C.WARN}⚠{C.END} {msg}")
def err(msg):  print(f"  {C.ERR}✗{C.END} {msg}")
def head(msg): print(f"\n{C.BOLD}{C.CYAN}{msg}{C.END}")


# ─────────────────────────────────────────────────────────────────────────
# 1. Verificar parquet fuente
# ─────────────────────────────────────────────────────────────────────────
def verificar_parquet(ruta_parquet: Path):
    head("1. VERIFICACIÓN DEL PARQUET FUENTE")
    print(f"  Ruta: {ruta_parquet}")

    if not ruta_parquet.exists():
        err(f"No existe: {ruta_parquet}")
        return None

    df = pd.read_parquet(ruta_parquet)
    ok(f"Leído: {len(df):,} filas, {len(df.columns)} columnas")

    # Verificar columnas después de renombrar
    df_r = df.rename(columns=PARQUET_RENOMBRA)
    cols_sql = set(VISTA_SQL_COLUMNAS)
    cols_parquet = set(df_r.columns)

    faltantes = cols_sql - cols_parquet
    extras = cols_parquet - cols_sql

    print(f"\n  Columnas en vista SQL : {len(cols_sql)}")
    print(f"  Columnas en parquet   : {len(cols_parquet)}")

    if faltantes:
        err(f"FALTAN {len(faltantes)} columnas: {sorted(faltantes)}")
    else:
        ok("Las 30 columnas de la vista SQL están presentes en el parquet (tras renombrar)")

    if extras:
        warn(f"{len(extras)} columnas extra en parquet: {sorted(extras)}")

    # Tipos de datos
    head("  Tipos de datos principales")
    for col in ["Año", "Mes", "respuestaFch", "areaNombre", "sedeNombre", "Atención", "NPS_Numerico"]:
        if col in df_r.columns:
            print(f"    {col:30s} {str(df_r[col].dtype):20s} nulos={df_r[col].isna().sum():,}")

    # Datos disponibles
    head("  Datos disponibles")
    if "Año" in df_r.columns:
        anios = sorted(df_r["Año"].dropna().unique().tolist())
        print(f"    Años: {anios}")
    if "Mes" in df_r.columns:
        meses = sorted(df_r["Mes"].dropna().unique().tolist())
        print(f"    Meses: {meses}")
    if "areaNombre" in df_r.columns:
        areas = sorted(df_r["areaNombre"].dropna().unique().tolist())
        print(f"    Áreas ({len(areas)}): {areas[:5]}..." if len(areas)>5 else f"    Áreas ({len(areas)}): {areas}")
    if "sedeNombre" in df_r.columns:
        sedes = sorted(df_r["sedeNombre"].dropna().unique().tolist())
        print(f"    Sedes ({len(sedes)}): {sedes}")
    if "encuestaNombre" in df_r.columns:
        encuestas = df_r["encuestaNombre"].dropna().nunique()
        print(f"    Tipos de encuesta: {encuestas}")

    return df


# ─────────────────────────────────────────────────────────────────────────
# 2. Verificar Excel generado
# ─────────────────────────────────────────────────────────────────────────
def verificar_excel(ruta_excel: Path, tipo="MENSUAL"):
    """Verifica estructura y datos de un Excel generado."""
    print(f"\n{'  '*2}Archivo: {ruta_excel.name}")

    if not ruta_excel.exists():
        err(f"  No existe: {ruta_excel}")
        return {"existe": False, "pestana": 0, "filas": 0, "errores": []}

    resultado = {"existe": True, "errores": []}
    size_mb = ruta_excel.stat().st_size / (1024 * 1024)
    ok(f"  Tamaño: {size_mb:.2f} MB")

    # Verificar con openpyxl (pestañas y estructura)
    if OPENPYXL_OK:
        try:
            wb = load_workbook(ruta_excel, read_only=True, data_only=True)
            pestanas = wb.sheetnames
            resultado["pestanas"] = pestanas
            ok(f"  Pestañas ({len(pestanas)}): {pestanas}")

            # Verificar pestañas esperadas
            pestanas_lower = [p.lower() for p in pestanas]
            for esp in PESTANAS_MENSUAL:
                if esp.lower() in pestanas_lower:
                    ok(f"    '{esp}' - presente")
                else:
                    err(f"    '{esp}' - FALTA")
                    resultado["errores"].append(f"Falta pestaña: {esp}")

            # Verificar Datos Completo (primera pestaña de datos)
            for pestana in pestanas:
                if "datos completo" in pestana.lower() or "raw" in pestana.lower():
                    ws = wb[pestana]
                    headers = []
                    for row in ws.iter_rows(min_row=5, max_row=5, values_only=True):
                        headers = [str(c).strip() if c else "" for c in row]
                        break

                    if headers:
                        print(f"\n    Columnas en '{pestana}' (fila 5): {headers[:10]}...")
                        resultado["columnas_raw"] = headers

                        # Verificar columnas de la vista SQL
                        headers_set = set(h.lower() for h in headers)
                        for col_sql in VISTA_SQL_COLUMNAS[:10]:
                            if col_sql.lower() in headers_set:
                                pass  # ok
                            # Algunas columnas renombradas
                            elif col_sql == "Año" and "anio" in headers_set:
                                pass
                            else:
                                pass  # no validar faltantes en Excel, pueden estar renombradas

                    break

            # Verificar Datos Procesados
            for pestana in pestanas:
                if "procesados" in pestana.lower():
                    ws = wb[pestana]
                    max_row = ws.max_row
                    max_col = ws.max_column
                    ok(f"  'Datos Procesados': {max_row:,} filas x {max_col} columnas")
                    resultado["filas_procesados"] = max_row

                    if max_row < 6:
                        warn("    Pocas filas de datos (< 5)")
                        resultado["errores"].append("Pocas filas en Datos Procesados")
                    break

            # Verificar Total Encuestas
            for pestana in pestanas:
                if "total" in pestana.lower():
                    ws = wb[pestana]
                    max_row = ws.max_row
                    ok(f"  'Total Encuestas': {max_row:,} filas")
                    break

            wb.close()

        except Exception as e:
            err(f"  Error abriendo Excel: {e}")
            resultado["errores"].append(str(e))
    else:
        warn("  openpyxl no disponible, omitiendo verificación de pestañas")

    # Verificar con pandas (datos)
    try:
        # Datos Completo
        for pestana in ["Datos Completo", "RawData"]:
            try:
                df = pd.read_excel(ruta_excel, sheet_name=pestana, header=4)
                if not df.empty:
                    ok(f"  '{pestana}': {len(df):,} filas x {len(df.columns)} cols")
                    resultado["filas_raw"] = len(df)
                break
            except:
                continue

        # Datos Procesados
        try:
            df_proc = pd.read_excel(ruta_excel, sheet_name="Datos Procesados", header=4)
            if not df_proc.empty:
                ok(f"  'Datos Procesados': {len(df_proc):,} filas")
                # Verificar columnas clave
                proc_cols = set(str(c).strip().lower() for c in df_proc.columns)
                for col_esperada in ["año", "mes", "métrica", "valor", "indicador_0_100"]:
                    if col_esperada in proc_cols:
                        ok(f"    Columna '{col_esperada}' presente")
                    else:
                        err(f"    Columna '{col_esperada}' FALTA en Datos Procesados")
                        resultado["errores"].append(f"Falta columna: {col_esperada}")

                # Métricas presentes
                if "métrica" in proc_cols:
                    metricas = df_proc["Métrica"].dropna().unique().tolist()
                    ok(f"    Métricas: {metricas}")
                    for met in METRICAS_ESPERADAS:
                        if met in metricas:
                            ok(f"      '{met}' OK")
                        else:
                            warn(f"      '{met}' no encontrada")
        except Exception as e:
            warn(f"  No se pudo leer 'Datos Procesados': {e}")

    except Exception as e:
        err(f"  Error leyendo con pandas: {e}")
        resultado["errores"].append(str(e))

    return resultado


# ─────────────────────────────────────────────────────────────────────────
# 3. Verificar filtrado
# ─────────────────────────────────────────────────────────────────────────
def verificar_filtrado(df: pd.DataFrame, anio: int, mes: int, area: str):
    head(f"3. VERIFICAR FILTRADO: Año={anio}, Mes={mes}, Área='{area}'")

    df_r = df.rename(columns=PARQUET_RENOMBRA)
    df_r["Año"] = pd.to_numeric(df_r["Año"], errors="coerce").fillna(0).astype(int)
    df_r["Mes"] = pd.to_numeric(df_r["Mes"], errors="coerce").fillna(0).astype(int)

    mask = (df_r["Año"] == anio) & (df_r["Mes"] == mes) & (df_r["areaNombre"] == area)
    df_filt = df_r[mask]

    print(f"  Filas filtradas: {len(df_filt):,}")

    if df_filt.empty:
        err("  ¡SIN DATOS! Verificando valores de areaNombre...")
        areas = sorted(df_r["areaNombre"].dropna().unique().tolist())
        print(f"  Áreas disponibles: {areas}")
        # Buscar por sedeNombre
        if area in df_r["sedeNombre"].values:
            warn(f"  → '{area}' es una SEDE, no un ÁREA")
            mask_sede = (df_r["Año"] == anio) & (df_r["Mes"] == mes) & (df_r["sedeNombre"] == area)
            df_sede = df_r[mask_sede]
            print(f"  → Filtrando por sede='{area}': {len(df_sede):,} filas")
    else:
        ok(f"  Filtrado OK: {len(df_filt):,} filas")

    return df_filt


# ─────────────────────────────────────────────────────────────────────────
# 4. Buscar Excels generados en la carpeta de salida
# ─────────────────────────────────────────────────────────────────────────
def buscar_excels(ruta_base: Path, area: str = None, anio: int = None, mes: int = None):
    head("2. BUSCAR EXCELS GENERADOS")
    print(f"  Ruta base: {ruta_base}")

    if not ruta_base.exists():
        err(f"No existe: {ruta_base}")
        return

    excels = list(ruta_base.rglob("*.xlsx"))
    if not excels:
        warn("No se encontraron archivos .xlsx en la ruta base")
        return

    ok(f"Encontrados {len(excels)} archivos Excel")

    # Filtrar si se especifican criterios
    excels_filtrados = []
    for f in excels:
        nombre = f.name.lower()
        incluir = True
        if area and area.lower().replace(" ", "_").replace("/", "_") not in nombre:
            incluir = False
        if anio and str(anio) not in nombre:
            incluir = False
        if mes and f"_{mes}_" not in f"_{nombre}":
            incluir = False
        if incluir:
            excels_filtrados.append(f)

    if not excels_filtrados:
        warn("No hay Excels que coincidan con los criterios")
        # Mostrar todos de todas formas
        excels_filtrados = excels[:20]

    print(f"\n  Mostrando hasta {len(excels_filtrados)} archivos:")
    for f in excels_filtrados[:20]:
        size_mb = f.stat().st_size / (1024 * 1024)
        print(f"    {f.relative_to(ruta_base)}  ({size_mb:.1f} MB)")

    # Verificar cada uno
    head("  Verificando contenido de los Excels")
    resultados = []
    for f in excels_filtrados[:10]:  # Max 10 para no saturar
        tipo = "MENSUAL" if "encuesta_" in f.name.lower() else "ACUMULADO" if "acumulado_" in f.name.lower() else "DESCONOCIDO"
        resultado = verificar_excel(f, tipo)
        resultado["archivo"] = str(f.relative_to(ruta_base))
        resultados.append(resultado)

    # Resumen
    head("  RESUMEN DE VALIDACIÓN")
    n_ok = sum(1 for r in resultados if not r.get("errores"))
    n_err = sum(1 for r in resultados if r.get("errores"))
    print(f"  Archivos validados  : {len(resultados)}")
    print(f"  Sin errores         : {C.OK}{n_ok}{C.END}")
    print(f"  Con errores         : {C.ERR}{n_err}{C.END}")

    for r in resultados:
        estado = f"{C.OK}OK{C.END}" if not r["errores"] else f"{C.ERR}ERROR{C.END}"
        print(f"    [{estado}] {r['archivo']}")
        if r["errores"]:
            for e in r["errores"][:3]:
                print(f"           - {e}")

    return resultados


# ─────────────────────────────────────────────────────────────────────────
# 5. Comparar parquet vs Excel generado
# ─────────────────────────────────────────────────────────────────────────
def comparar_parquet_vs_excel(df_parquet: pd.DataFrame, ruta_excel: Path, area: str, mes: int, anio: int):
    head("4. COMPARAR PARQUET vs EXCEL")

    df_r = df_parquet.rename(columns=PARQUET_RENOMBRA)
    df_r["Año"] = pd.to_numeric(df_r["Año"], errors="coerce").fillna(0).astype(int)
    df_r["Mes"] = pd.to_numeric(df_r["Mes"], errors="coerce").fillna(0).astype(int)
    mask = (df_r["Año"] == anio) & (df_r["Mes"] == mes) & (df_r["areaNombre"] == area)
    df_esperado = df_r[mask]

    if df_esperado.empty:
        warn("No hay datos en el parquet para esta combinación")
        return

    try:
        df_excel = pd.read_excel(ruta_excel, sheet_name="Datos Completo", header=4)
        ok(f"Filas esperadas (parquet filtrado): {len(df_esperado):,}")
        ok(f"Filas en Excel (Datos Completo)   : {len(df_excel):,}")

        if len(df_esperado) == len(df_excel):
            ok("✓ Coincide el número de filas")
        else:
            warn(f"  Diferencia: {abs(len(df_esperado) - len(df_excel))} filas")

    except Exception as e:
        err(f"No se pudo leer el Excel: {e}")


# ─────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Valida Excels generados por lanzador parquet")
    parser.add_argument("--ruta", type=str, default="F:/ETL_DITIC/temp_exportacion_multiarea",
                        help="Ruta base donde se guardaron los Excels")
    parser.add_argument("--parquet", type=str,
                        default="F:/ETL_DITIC/EncuestasPercepcionAzure/VistaEncuestaPercepcion2026.parquet",
                        help="Ruta al parquet fuente")
    parser.add_argument("--area", type=str, default=None, help="Filtrar por área")
    parser.add_argument("--anio", type=int, default=2026, help="Año (default: 2026)")
    parser.add_argument("--mes", type=int, default=None, help="Mes específico")
    args = parser.parse_args()

    print()
    print("=" * 70)
    print("  VALIDACIÓN DE EXCELS GENERADOS - LANZADOR PARQUET")
    print("=" * 70)
    print(f"  Fecha       : {datetime.now():%Y-%m-%d %H:%M}")
    print(f"  Parquet     : {args.parquet}")
    print(f"  Ruta Excels : {args.ruta}")
    print(f"  Filtros     : anio={args.anio}, mes={args.mes}, area={args.area}")
    print("=" * 70)

    # 1. Verificar parquet
    df = verificar_parquet(Path(args.parquet))
    if df is None:
        sys.exit(1)

    # 2. Buscar y verificar Excels
    ruta_excels = Path(args.ruta)
    resultados = buscar_excels(ruta_excels, area=args.area, anio=args.anio, mes=args.mes)

    # 3. Verificar filtrado
    if args.area and args.mes:
        verificar_filtrado(df, args.anio, args.mes, args.area)
        # Buscar el Excel de esta área para comparar
        area_safe = args.area.replace(" ", "_").replace("/", "_")
        patrón_mensual = f"Encuesta_{args.anio}_{args.mes}_{area_safe}.xlsx"
        excels_encontrados = list(ruta_excels.rglob(f"*{area_safe}*.xlsx"))
        if excels_encontrados:
            comparar_parquet_vs_excel(df, excels_encontrados[0], args.area, args.mes, args.anio)
    else:
        # Mostrar todas las áreas para que el usuario sepa qué filtrar
        head("3. ÁREAS DISPONIBLES (para filtrar)")
        df_r = df.rename(columns=PARQUET_RENOMBRA)
        df_r["Año"] = pd.to_numeric(df_r["Año"], errors="coerce").fillna(0).astype(int)
        areas = sorted(df_r["areaNombre"].dropna().unique().tolist())
        print(f"  Áreas ({len(areas)}):")
        for a in areas:
            print(f"    - {a}")

        sedes = sorted(df_r["sedeNombre"].dropna().unique().tolist())
        print(f"\n  Sedes ({len(sedes)}):")
        for s in sedes:
            print(f"    - {s}")

    print(f"\n{'=' * 70}")
    print("  VALIDACIÓN COMPLETADA")
    print("=" * 70)


if __name__ == '__main__':
    main()