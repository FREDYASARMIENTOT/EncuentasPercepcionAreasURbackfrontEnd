"""
orquestador_parquet.py - Módulo backend para lectura de encuestas desde PARQUET

Reemplaza la consulta SQL a View_respuestas_encuesta_percepcion_historica_optimizada
por la lectura y filtrado del archivo parquet.

Uso desde el portal:
    from backend.orquestador_parquet import ParquetLauncher
    
    launcher = ParquetLauncher()
    areas = launcher.get_areas_disponibles(2026)
    meses = launcher.get_meses_disponibles(2026)
    datos = launcher.filtrar_datos(anio=2026, mes=5, area="CRAI")
"""

import os
import sys
import gc
import logging
import traceback
from pathlib import Path
from io import BytesIO
from datetime import datetime
from typing import Optional, List, Dict, Any
from collections import defaultdict
from contextlib import contextmanager
import time
import warnings

import pandas as pd
import numpy as np

from .database import SessionLocal
from .models import Area, CargaArea, JobEjecucion, JobParametro

# Suprimir warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
warnings.filterwarnings('ignore', category=FutureWarning)

# Azure Blob Storage
try:
    from azure.storage.blob import BlobServiceClient
    AZURE_AVAILABLE = True
except ImportError:
    AZURE_AVAILABLE = False

logger = logging.getLogger(__name__)

# --------------------------------------------------------------------------
# Configuración
# --------------------------------------------------------------------------
ROOT_DIR = Path(__file__).resolve().parents[1]
PARQUET_LOCAL_PATH = ROOT_DIR / "VistaEncuestaPercepcion2026.parquet"

# Mapeo de columnas parquet -> nombres vista SQL
PARQUET_COLUMN_MAP = {
    "Anio": "Año",
    "Atencion": "Atención",
    "Comunicacion_y_acceso": "Comunicación y acceso",
    "Resolucion_de_la_necesidad": "Resolución de la necesidad",
    "Tiempo_de_respuesta": "Tiempo de respuesta",
}

# Métricas para melt
METRIC_COLUMNS = ['Atención', 'Comunicación y acceso', 'Eficiencia', 'NPS',
                  'Resolución de la necesidad', 'Tiempo de respuesta']

# --------------------------------------------------------------------------
# Performance Monitor
# --------------------------------------------------------------------------
class PerformanceMonitor:
    def __init__(self):
        self.metrics = defaultdict(lambda: {'count': 0, 'total_time': 0, 'errors': 0})
    
    @contextmanager
    def measure(self, stage_name):
        start_time = time.time()
        try:
            yield
            elapsed = time.time() - start_time
            self.metrics[stage_name]['count'] += 1
            self.metrics[stage_name]['total_time'] += elapsed
            logger.info(f"[{stage_name}] Completado en {elapsed:.2f}s")
        except Exception as e:
            self.metrics[stage_name]['errors'] += 1
            logger.error(f"[{stage_name}] Error: {e}")
            raise
    
    def get_summary(self):
        lines = []
        for stage, data in sorted(self.metrics.items()):
            lines.append(f"  {stage}: {data['total_time']:.2f}s | Errores {data['errors']}")
        return "\n".join(lines)


# --------------------------------------------------------------------------
# Lector de Parquet (Singleton con cache)
# --------------------------------------------------------------------------
class ParquetReader:
    """Lee y cachea el parquet desde local o Azure Blob Storage."""
    
    _cached_df: Optional[pd.DataFrame] = None
    _cached_source: Optional[str] = None
    _loaded_at: Optional[datetime] = None
    
    @classmethod
    def get_dataframe(cls, force_reload: bool = False) -> pd.DataFrame:
        """Obtiene el DataFrame del parquet, cacheándolo para uso posterior."""
        if cls._cached_df is not None and not force_reload:
            return cls._cached_df
        
        # Intentar primero desde local
        if PARQUET_LOCAL_PATH.exists() and not force_reload:
            logger.info(f"Leyendo parquet LOCAL: {PARQUET_LOCAL_PATH}")
            try:
                cls._cached_df = pd.read_parquet(PARQUET_LOCAL_PATH)
                cls._cached_source = "local"
                cls._loaded_at = datetime.now()
                logger.info(f"Parquet cargado: {len(cls._cached_df):,} filas | {len(cls._cached_df.columns)} columnas")
                return cls._cached_df
            except Exception as e:
                logger.warning(f"Error leyendo parquet local: {e}")
        
        # Descargar desde Azure Blob
        if AZURE_AVAILABLE:
            logger.info("Descargando parquet desde Azure Blob...")
            try:
                az_url = os.getenv('AZURE_STORAGE_ACCOUNT_URL', 'https://saurdatamining.blob.core.windows.net')
                az_container = os.getenv('AZURE_CONTAINER_NAME', 'fs-encuestaspercepcion')
                az_sas = os.getenv('AZURE_SAS_TOKEN', '')
                
                url = az_url.rstrip("/")
                sas = az_sas.lstrip("?")
                blob_url = f"{url}?{sas}" if sas else url
                
                blob_service = BlobServiceClient(blob_url)
                blob_client = blob_service.get_blob_client(
                    container=az_container,
                    blob="EncuestasPercepcion/VistaEncuestaPercepcion2026.parquet"
                )
                raw = blob_client.download_blob().readall()
                cls._cached_df = pd.read_parquet(BytesIO(raw))
                cls._cached_source = "azure"
                cls._loaded_at = datetime.now()
                logger.info(f"Parquet descargado: {len(raw):,} bytes | {len(cls._cached_df):,} filas")
                
                # Guardar en local para futuras consultas
                try:
                    cls._cached_df.to_parquet(PARQUET_LOCAL_PATH, index=False, engine="pyarrow")
                    logger.info(f"Cached localmente en: {PARQUET_LOCAL_PATH}")
                except Exception as e:
                    logger.warning(f"No se pudo cachear localmente: {e}")
                
                return cls._cached_df
            except Exception as e:
                logger.error(f"Error descargando de Azure: {e}")
        
        raise FileNotFoundError(f"No se pudo cargar el parquet")
    
    @classmethod
    def reload(cls):
        """Fuerza la recarga del parquet."""
        cls._cached_df = None
        cls._cached_source = None
        cls._loaded_at = None
        return cls.get_dataframe(force_reload=True)
    
    @classmethod
    def get_info(cls) -> Dict[str, Any]:
        """Obtiene información básica del parquet cargado."""
        df = cls.get_dataframe()
        return {
            "source": cls._cached_source,
            "loaded_at": cls._loaded_at.isoformat() if cls._loaded_at else None,
            "rows": len(df),
            "columns": len(df.columns),
            "local_path": str(PARQUET_LOCAL_PATH),
            "local_exists": PARQUET_LOCAL_PATH.exists(),
            "azure_available": AZURE_AVAILABLE,
        }
    
    @classmethod
    def _normalizar_columna(cls, df: pd.DataFrame, nombre_buscado: str) -> Optional[str]:
        """Busca el nombre real de la columna."""
        if nombre_buscado in df.columns:
            return nombre_buscado
        variantes = {
            'Año': ['Anio', 'año', 'anio'],
            'Mes': ['mes'],
            'areaNombre': ['AreaNombre', 'area_nombre'],
        }
        for variante in variantes.get(nombre_buscado, []):
            if variante in df.columns:
                return variante
        return None
    
    @classmethod
    def get_areas_disponibles(cls, anio: int) -> List[str]:
        """Obtiene las áreas disponibles para un año específico."""
        df = cls.get_dataframe()
        df_ren = df.rename(columns=PARQUET_COLUMN_MAP)
        col_anio = cls._normalizar_columna(df_ren, 'Año')
        col_area = cls._normalizar_columna(df_ren, 'areaNombre')
        
        if not col_anio or not col_area:
            return []
        
        df_ren[col_anio] = pd.to_numeric(df_ren[col_anio], errors='coerce').fillna(0).astype(int)
        areas = df_ren[df_ren[col_anio] == anio][col_area].dropna().unique().tolist()
        return sorted([a for a in areas if a and str(a).strip()])
    
    @classmethod
    def get_meses_disponibles(cls, anio: int) -> List[int]:
        """Obtiene los meses disponibles para un año específico."""
        df = cls.get_dataframe()
        df_ren = df.rename(columns=PARQUET_COLUMN_MAP)
        col_anio = cls._normalizar_columna(df_ren, 'Año')
        col_mes = cls._normalizar_columna(df_ren, 'Mes')
        
        if not col_anio or not col_mes:
            return []
        
        df_ren[col_anio] = pd.to_numeric(df_ren[col_anio], errors='coerce').fillna(0).astype(int)
        meses = df_ren[df_ren[col_anio] == anio][col_mes].dropna().unique().tolist()
        return sorted([int(m) for m in meses if m is not None])
    
    @classmethod
    def get_encuestas_disponibles(cls, anio: int) -> List[str]:
        """Obtiene los nombres de encuestas disponibles para un año."""
        df = cls.get_dataframe()
        col_anio = cls._normalizar_columna(df, 'Año')
        if not col_anio:
            return []
        df[col_anio] = pd.to_numeric(df[col_anio], errors='coerce').fillna(0).astype(int)
        encuestas = df[df[col_anio] == anio]['encuestaNombre'].dropna().unique().tolist()
        return sorted([e for e in encuestas if e and str(e).strip()])
    
    @classmethod
    def get_sedes_disponibles(cls, anio: int) -> List[str]:
        """Obtiene las sedes disponibles para un año."""
        df = cls.get_dataframe()
        col_anio = cls._normalizar_columna(df, 'Año')
        if not col_anio:
            return []
        df[col_anio] = pd.to_numeric(df[col_anio], errors='coerce').fillna(0).astype(int)
        sedes = df[df[col_anio] == anio]['sedeNombre'].dropna().unique().tolist()
        return sorted([s for s in sedes if s and str(s).strip()])
    
    @classmethod
    def filtrar_datos(cls, anio: int, mes: int, area: str) -> pd.DataFrame:
        """Filtra los datos del parquet por Año, Mes y área.Nombre."""
        df = cls.get_dataframe()
        
        # Renombrar columnas para coincidir con la vista SQL
        df_renombrado = df.rename(columns=PARQUET_COLUMN_MAP)
        
        col_anio = 'Año' if 'Año' in df_renombrado.columns else cls._normalizar_columna(df_renombrado, 'Año')
        col_mes = 'Mes' if 'Mes' in df_renombrado.columns else cls._normalizar_columna(df_renombrado, 'Mes')
        col_area = 'areaNombre' if 'areaNombre' in df_renombrado.columns else cls._normalizar_columna(df_renombrado, 'areaNombre')
        
        if not col_anio or not col_mes or not col_area:
            logger.error(f"Columnas no encontradas. Disponibles: {list(df_renombrado.columns)}")
            return pd.DataFrame()
        
        df_renombrado[col_anio] = pd.to_numeric(df_renombrado[col_anio], errors='coerce').fillna(0).astype(int)
        df_renombrado[col_mes] = pd.to_numeric(df_renombrado[col_mes], errors='coerce').fillna(0).astype(int)
        
        mask = (df_renombrado[col_anio] == anio) & (df_renombrado[col_mes] == mes) & (df_renombrado[col_area] == area)
        df_filtrado = df_renombrado[mask].copy()
        
        logger.info(f"Filtrado: {len(df_filtrado):,} filas para {area} | {mes}/{anio}")
        return df_filtrado
    
    @classmethod
    def filtrar_datos_avanzado(cls, anio: int = None, mes: int = None, area: str = None,
                                sede: str = None, servicio: str = None, encuesta: str = None) -> pd.DataFrame:
        """Filtrado avanzado con múltiples criterios opcionales."""
        df = cls.get_dataframe()
        df_r = df.rename(columns=PARQUET_COLUMN_MAP)
        
        filters = []
        
        if anio is not None:
            col = 'Año' if 'Año' in df_r.columns else cls._normalizar_columna(df_r, 'Año')
            if col:
                df_r[col] = pd.to_numeric(df_r[col], errors='coerce').fillna(0).astype(int)
                filters.append(df_r[col] == anio)
        
        if mes is not None:
            col = 'Mes' if 'Mes' in df_r.columns else cls._normalizar_columna(df_r, 'Mes')
            if col:
                df_r[col] = pd.to_numeric(df_r[col], errors='coerce').fillna(0).astype(int)
                filters.append(df_r[col] == mes)
        
        if area is not None:
            col = 'areaNombre' if 'areaNombre' in df_r.columns else cls._normalizar_columna(df_r, 'areaNombre')
            if col:
                filters.append(df_r[col] == area)
        
        if sede is not None:
            col = 'sedeNombre' if 'sedeNombre' in df_r.columns else None
            if col:
                filters.append(df_r[col] == sede)
        
        if servicio is not None:
            col = 'servicioNombre' if 'servicioNombre' in df_r.columns else None
            if col:
                filters.append(df_r[col] == servicio)
        
        if encuesta is not None:
            col = 'encuestaNombre' if 'encuestaNombre' in df_r.columns else None
            if col:
                filters.append(df_r[col] == encuesta)
        
        if not filters:
            return df_r.copy()
        
        mask = filters[0]
        for f in filters[1:]:
            mask = mask & f
        
        return df_r[mask].copy()
    
    @classmethod
    def get_resumen(cls, anio: int, mes: int = None, area: str = None) -> Dict[str, Any]:
        """Obtiene un resumen estadístico de los datos filtrados."""
        df = cls.filtrar_datos_avanzado(anio=anio, mes=mes, area=area)
        
        if df.empty:
            return {"rows": 0, "encuestados": 0, "encuestas": 0}
        
        return {
            "rows": len(df),
            "encuestados": int(df['encuestadoId'].nunique()) if 'encuestadoId' in df.columns else 0,
            "consecutivos": int(df['consecutivo'].nunique()) if 'consecutivo' in df.columns else 0,
            "encuestas": int(df['encuestaNombre'].nunique()) if 'encuestaNombre' in df.columns else 0,
            "sedes": sorted(df['sedeNombre'].dropna().unique().tolist()) if 'sedeNombre' in df.columns else [],
            "servicios": sorted(df['servicioNombre'].dropna().unique().tolist()) if 'servicioNombre' in df.columns else [],
        }


# --------------------------------------------------------------------------
# Data Engine (Procesamiento de indicadores)
# --------------------------------------------------------------------------
def scalarize(x):
    return x[0] if isinstance(x, (list, np.ndarray)) and len(x) > 0 else x

def calcular_indicador(valor):
    """Calcula indicador 0-100 a partir de un valor."""
    if pd.isna(valor): return None
    s = str(valor).strip().upper()
    
    if 'EXCELENTE' in s or 'MUY BUENO' in s or 'MUY BIEN' in s: return 100.0
    if 'MUY MALO' in s or 'PÉSIMO' in s or 'NUNCA' in s: return 0.0
    if s in ['SI', 'YES', '1']: return 100.0
    if s in ['NO', '0']: return 0.0
    try:
        n = float(s.replace(',', '.'))
        if 1 <= n <= 5: return {1: 0.0, 2: 25.0, 3: 50.0, 4: 75.0, 5: 100.0}.get(round(n), (n-1)*25.0)
        if 5 < n <= 10: return (n-1)*(100.0/9.0)
        if 10 < n <= 100: return n
    except: pass
    return 50.0

def fusionar_metricas(df):
    """Fusiona métricas numéricas con sus equivalentes de texto."""
    if 'Métrica' in df.columns:
        df['Métrica'] = df['Métrica'].replace({'NPS_Numerico': 'NPS', 'Atencion_Numerica': 'Atención'})
    else:
        for old, new in {'NPS_Numerico': 'NPS', 'Atencion_Numerica': 'Atención'}.items():
            if old in df.columns and new in df.columns:
                df[new] = df[new].fillna(df[old])
                df.drop(columns=[old], inplace=True)
            elif old in df.columns:
                df.rename(columns={old: new}, inplace=True)
    return df

def procesar_dataframe_melt(df_raw: pd.DataFrame) -> pd.DataFrame:
    """Procesa los datos raw aplicando melt y calculando indicadores."""
    df = df_raw.copy()
    
    # Columnas métricas adicionales
    extra_mets = ['preguntaSinIndicador', 'NPS_Numerico', 'Atencion_Numerica']
    all_mets = METRIC_COLUMNS + extra_mets
    
    fij = [c for c in df.columns if c not in all_mets]
    
    # Melt
    melt = df.melt(
        id_vars=fij,
        value_vars=[c for c in all_mets if c in df.columns],
        var_name='Métrica',
        value_name='Valor'
    ).dropna(subset=['Valor'])
    
    # Procesar
    melt = fusionar_metricas(melt)
    if 'respuestaId' in melt.columns:
        melt.sort_values(by=['respuestaId', 'Métrica', 'Valor'], na_position='first', inplace=True)
        melt.drop_duplicates(subset=['respuestaId', 'Métrica'], keep='last', inplace=True)
    
    melt['Valor'] = melt['Valor'].apply(scalarize)
    melt['Indicador_0_100'] = melt['Valor'].apply(calcular_indicador)
    
    for c in ['Año', 'Mes']:
        if c in melt.columns:
            melt[c] = pd.to_numeric(melt[c], errors='coerce').fillna(0).astype(int)
    
    return melt


# --------------------------------------------------------------------------
# Lanzador Parquet (Integración con Portal)
# --------------------------------------------------------------------------
class ParquetLauncher:
    """
    Integración del lanzador parquet con el portal web.
    Proporciona los mismos métodos que el lanzador SQL pero usando parquet.
    """
    
    def __init__(self):
        self.reader = ParquetReader()
        self.perf_monitor = PerformanceMonitor()
    
    def get_parquet_info(self) -> Dict[str, Any]:
        """Información del parquet."""
        return self.reader.get_info()
    
    def reload_parquet(self) -> Dict[str, Any]:
        """Recargar el parquet."""
        self.reader.reload()
        return self.get_parquet_info()
    
    def get_areas_disponibles(self, anio: int) -> List[str]:
        """Obtener áreas disponibles para un año."""
        return self.reader.get_areas_disponibles(anio)
    
    def get_meses_disponibles(self, anio: int) -> List[int]:
        """Obtener meses disponibles para un año."""
        return self.reader.get_meses_disponibles(anio)
    
    def get_resumen_datos(self, anio: int, mes: int = None, area: str = None) -> Dict[str, Any]:
        """Obtener resumen estadístico de los datos."""
        return self.reader.get_resumen(anio, mes, area)
    
    def obtener_datos_filtrados(self, anio: int, mes: int = None, area: str = None,
                                 sede: str = None, servicio: str = None) -> pd.DataFrame:
        """Obtener datos filtrados con múltiples criterios."""
        return self.reader.filtrar_datos_avanzado(
            anio=anio, mes=mes, area=area,
            sede=sede, servicio=servicio
        )
    
    def procesar_y_exportar(self, anio: int, mes: int, area: str,
                             output_dir: Optional[str] = None) -> Dict[str, Any]:
        """
        Procesa los datos filtrados y genera archivos Excel.
        Versión simplificada que retorna datos procesados y estadísticas.
        """
        result = {
            "success": False,
            "area": area,
            "anio": anio,
            "mes": mes,
            "rows_raw": 0,
            "rows_processed": 0,
            "encuestados": 0,
            "encuestas": 0,
            "files": [],
            "error": None,
            "duration_seconds": 0,
        }
        
        start = time.time()
        
        try:
            # 1. Obtener datos filtrados
            with self.perf_monitor.measure("extraccion_parquet"):
                raw = self.reader.filtrar_datos(anio, mes, area)
            
            if raw.empty:
                result["error"] = f"Sin datos para {area} en {mes}/{anio}"
                return result
            
            result["rows_raw"] = len(raw)
            
            # 2. Procesar datos
            with self.perf_monitor.measure("procesamiento"):
                processed = procesar_dataframe_melt(raw)
            
            result["rows_processed"] = len(processed)
            result["encuestados"] = int(processed['encuestadoId'].nunique()) if 'encuestadoId' in processed.columns else 0
            result["encuestas"] = int(processed['consecutivo'].nunique()) if 'consecutivo' in processed.columns else 0
            
            # 3. Generar archivos si se indica directorio
            if output_dir:
                with self.perf_monitor.measure("generacion_excel"):
                    from openpyxl import Workbook
                    from openpyxl.utils.dataframe import dataframe_to_rows
                    from openpyxl.utils import get_column_letter
                    from openpyxl.styles import Font, PatternFill
                    
                    output_path = Path(output_dir)
                    output_path.mkdir(parents=True, exist_ok=True)
                    
                    s_area = area.replace('/', '_').replace(' ', '_')
                    
                    # Excel Mensual
                    f_mensual = output_path / f"Encuesta_{anio}_{mes}_{s_area}.xlsx"
                    wb = Workbook()
                    
                    # Datos Completo
                    ws = wb.active
                    ws.title = "Datos Completo"
                    for r in dataframe_to_rows(raw, index=False, header=True):
                        ws.append(list(r))
                    # Ajustar anchos
                    for col in ws.columns:
                        max_length = 0
                        col_letter = get_column_letter(col[0].column)
                        for cell in col:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)
                    
                    # Datos Procesados
                    ws2 = wb.create_sheet("Datos Procesados")
                    proc_cols = ['encuestadoId', 'consecutivo', 'Año', 'Mes', 'Métrica', 'Valor', 'Indicador_0_100']
                    proc_cols = [c for c in proc_cols if c in processed.columns]
                    ws2.append(proc_cols)
                    for _, row in processed[proc_cols].iterrows():
                        ws2.append(list(row))
                    
                    # Resumen
                    ws3 = wb.create_sheet("Resumen")
                    ws3.append(["Métrica", "Cantidad", "Promedio Indicador"])
                    resumen = processed.groupby('Métrica').agg(
                        Cantidad=('respuestaId', 'nunique') if 'respuestaId' in processed.columns else ('Valor', 'count'),
                        Promedio=('Indicador_0_100', 'mean')
                    ).reset_index()
                    for _, row in resumen.iterrows():
                        ws3.append(list(row))
                    
                    wb.save(f_mensual)
                    result["files"].append(str(f_mensual))
                    
                    # Excel Acumulado
                    f_acum = output_path / f"Acumulado_{anio}_{s_area}.xlsx"
                    wb2 = Workbook()
                    ws_all = wb2.active
                    ws_all.title = "Datos Procesados"
                    ws_all.append(proc_cols)
                    for _, row in processed[proc_cols].iterrows():
                        ws_all.append(list(row))
                    wb2.save(f_acum)
                    result["files"].append(str(f_acum))
            
            result["success"] = True
            logger.info(f"Procesamiento completado: {area} | {mes}/{anio} | {result['rows_raw']:,} filas")
            
        except Exception as e:
            result["error"] = f"{e}\n{traceback.format_exc()}"
            logger.error(f"Error procesando {area} | {mes}/{anio}: {e}")
        finally:
            result["duration_seconds"] = time.time() - start
            gc.collect()
        
        return result
    
    def ejecutar_para_combinacion(self, anio: int, mes: int, area: str,
                                    area_id: Optional[int] = None,
                                    tipo_carga: str = "mensual",
                                    usuario: str = "portal") -> Dict[str, Any]:
        """
        Ejecuta el procesamiento completo para una combinación anio/mes/area,
        registrando en la base de datos del portal.
        """
        db = SessionLocal()
        result = {"success": False, "error": None, "carga_id": None}
        
        try:
            # Crear registro de carga
            carga = CargaArea(
                Mes=mes or 0,
                Anio=anio,
                AreaId=area_id,
                TipoCarga=tipo_carga,
                Estado="En proceso",
                Observaciones=f"Lanzador Parquet - {area} | {mes}/{anio}",
            )
            db.add(carga)
            db.flush()
            result["carga_id"] = carga.CargaId
            
            # Ejecutar procesamiento
            output_dir = os.getenv("BASE_CARPETA", "F:/ETL_DITIC/temp_exportacion_multiarea")
            proc_result = self.procesar_y_exportar(anio, mes, area, output_dir=output_dir)
            
            if proc_result["success"]:
                carga.Estado = "Exitoso"
                carga.RegistrosProcesados = proc_result["rows_processed"]
                carga.Observaciones = f"Parquet OK: {proc_result['rows_raw']:,} raw -> {proc_result['rows_processed']:,} procesados | {proc_result['encuestados']:,} encuestados | {proc_result['duration_seconds']:.1f}s"
            else:
                carga.Estado = "Error"
                carga.Observaciones = f"Error Parquet: {proc_result['error']}"
            
            db.commit()
            result["success"] = proc_result["success"]
            result["stats"] = proc_result
            
        except Exception as e:
            db.rollback()
            result["error"] = str(e)
            logger.error(f"Error en ejecutar_para_combinacion: {e}")
        finally:
            db.close()
        
        return result


# --------------------------------------------------------------------------
# Funciones de conveniencia
# --------------------------------------------------------------------------
_parquet_launcher_instance = None

def get_parquet_launcher() -> ParquetLauncher:
    """Obtiene la instancia singleton del lanzador parquet."""
    global _parquet_launcher_instance
    if _parquet_launcher_instance is None:
        _parquet_launcher_instance = ParquetLauncher()
    return _parquet_launcher_instance

def get_parquet_status() -> Dict[str, Any]:
    """Obtiene el estado del parquet y datos disponibles."""
    launcher = get_parquet_launcher()
    info = launcher.get_parquet_info()
    
    # Añadir años y meses disponibles
    if info["rows"] > 0:
        df = ParquetReader.get_dataframe()
        df_r = df.rename(columns=PARQUET_COLUMN_MAP)
        col_anio = ParquetReader._normalizar_columna(df_r, 'Año')
        if col_anio:
            df_r[col_anio] = pd.to_numeric(df_r[col_anio], errors='coerce').fillna(0).astype(int)
            info["anios_disponibles"] = sorted(df_r[col_anio].unique().tolist())
    
    return info